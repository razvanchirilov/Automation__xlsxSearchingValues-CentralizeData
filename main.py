import os
import openpyxl
from openpyxl.styles import Font

def search_and_extract(folder_path, new_file_path):
    # Create a new Excel file to store the extracted information
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    
    # Add the column titles
    new_ws.append(["Client", "Anexa", "Data Eveniment", "Denumire Proiect", "Brand Facturat", "Rebate(EURO)"])
    
    # Set the dimension of the column, the size of font and set it to bold
    width = 20
    bold_font = Font(bold=True, size=13)
    for col in "ABCDEFGH":
        new_ws.column_dimensions[col].width = width
        for row in new_ws.iter_rows(min_col=ord(col)-ord('A')+1, max_col=ord(col)-ord('A')+1, max_row=new_ws.max_row):
            for cell in row:
                cell.font = bold_font
        
    # Iterate through all files in the folder
    for file_name in os.listdir(folder_path):
        # Check if file is an Excel file (ends with .xls or .xlsx)
        if file_name.endswith(".xls") or file_name.endswith(".xlsx"):
            # Open the Excel file
            wb = openpyxl.load_workbook(os.path.join(folder_path, file_name))
            ws = wb.active         

            # Extract the values from the specific cells
            value1 = ws["B3"].value
            value2 = ws["B4"].value
            value3 = ws["B5"].value
            value4 = ws["B6"].value
            value5 = ws["B7"].value
            value6 = ws["B8"].value         
      
            # Write the values to the new Excel file
            new_ws.append([value1, value2, value3, value4, value5, value6])   
              
    # Save the new Excel file
    new_wb.save(new_file_path)


search_and_extract("CE", "centralization.xlsx")